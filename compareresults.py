#!/bin/sh
# -*- coding: utf-8 -*-

'''
This program compares two excel workbooks and generates a third workbook
with discrepancies highlighted. The purpose is to facilitate the file comparison
process for a comprehensive capital analysis review (CCAR) initiative during a
summer internship.
'''

import openpyxl
print sys.getdefaultencoding()

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# Load input files from the same directory
df1 = load_workbook(filename = 'Results1.xlsx')
df2 = load_workbook(filename = 'Results2.xlsx')

# Generate a third file that compares the two input files
df3 = openpyxl.Workbook()

# Matches cells with worksheets
df11 = df1.worksheets[0]
df22 = df2.worksheets[0]
df33 = df3.worksheets[0]

# Conditional coloring. green for equal values, red of different values
redFill = PatternFill(start_color='FFC7CE',
                   end_color='FFC7CE',
                   fill_type='solid')
greenFill = PatternFill(start_color='C6EFCE',
                   end_color='C6EFCE',
                   fill_type='solid')   

# Copies values of the first row of the first input file to the generated report
for x in range(1, df11.max_row + 1):
    df3_row = df33.cell(row = x, column = 1)
    df1_row = df11.cell(row = x, column = 1)
    df3_row.value = df1_row.value
    
# Copies values of the first column of the first input file to the generated report
for y in range(1, df11.max_column + 1):
    df3_col = df33.cell(row = 1, column = y)
    df1_col = df11.cell(row = 1, column = y)
    df3_col.value = df1_col.value

# Compares cells from first input file with cells from second input file
for i in range(2, df11.max_row + 1): # Iterate through rows
    
    for j in range(2, df11.max_column + 1): # Iterate through columns
        df1_cell = df11.cell(row = i, column = j)
        df2_cell = df22.cell(row = i, column = j)
        df3_cell = df33.cell(row = i, column = j)
        
        if df1_cell.value == df2_cell.value:
            df3_cell.value = 'TRUE'
            df3_cell.font = Font(color='006100')
            df3_cell.fill = greenFill
           
        elif df1_cell.value != df2_cell.value:
            df3_cell.value = 'FALSE'
            df3_cell.font = Font(color='9C0006')
            df3_cell.fill = redFill    
    
    # Save results to a generated excel file called 'finalresults.xlsx' in the same directory
    df3.save('Resultscompared.xlsx')
